import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class ExcelFormatter:
    def format_and_save(self, data, file_path):
        df = pd.DataFrame(data)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')

        # Convert the dataframe to an Excel object
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        worksheet = writer.sheets['Sheet1']

        self._format_worksheet(worksheet)

        # Save the Excel file
        writer.close()
        print(f"Model evaluations exported to {file_path}")

    def _format_worksheet(self, worksheet):
        # Formatting definitions
        header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        cell_font = Font(name='Calibri', size=11)
        center_aligned_text = Alignment(horizontal='center')
        border_side = Side(border_style="thin")
        square_border = Border(top=border_side, right=border_side, bottom=border_side, left=border_side)
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

        # Apply formatting to header
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.alignment = center_aligned_text
            cell.border = square_border
            cell.fill = header_fill

        # Set the width of the columns
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length

        # Apply formatting to each cell
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font
                cell.alignment = center_aligned_text
                cell.border = square_border
